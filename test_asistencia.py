import unittest
import os
import csv
from datetime import datetime
from openpyxl import load_workbook

# Importar las funciones a probar desde el archivo de código
from generarListado_AlvaroMartinez import generar_hoja_asistencia, convertir_excel_a_pdf

class TestAsistencia(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        """Crear un archivo CSV de prueba antes de ejecutar las pruebas."""
        cls.csv_file = "participantes_ficticios.csv"
        cls.excel_file = "hoja_asistencia.xlsx"
        cls.pdf_file = "hoja_asistencia.pdf"
        
        # Crear un CSV de prueba
        datos = [
            {"Nombre": "Juan", "Apellido(s)": "García"},
            {"Nombre": "María", "Apellido(s)": "López"},
            {"Nombre": "Carlos", "Apellido(s)": "Martínez"},
            {"Nombre": "Ana", "Apellido(s)": "Fernández"},
            {"Nombre": "Luis", "Apellido(s)": "Pérez"},
            {"Nombre": "Laura", "Apellido(s)": "Gómez"},
            {"Nombre": "Pedro", "Apellido(s)": "Rodríguez"},
            {"Nombre": "Sofía", "Apellido(s)": "Díaz"},
            {"Nombre": "Miguel", "Apellido(s)": "Sánchez"},
            {"Nombre": "Elena", "Apellido(s)": "Romero"},
        ]
        with open(cls.csv_file, mode="w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=["Nombre", "Apellido(s)"])
            writer.writeheader()
            writer.writerows(datos)

    @classmethod

    def test_generar_hoja_asistencia(self):
        """Prueba que la función generar_hoja_asistencia crea un archivo Excel correctamente."""
        # Ejecutar la función
        generar_hoja_asistencia(
            self.csv_file, self.excel_file, "PFIS", "Semana 2", "Tema_AlvaroMartinez"
        )
        
        # Verificar que el archivo Excel se ha creado
        self.assertTrue(os.path.exists(self.excel_file), "El archivo Excel no se ha creado")
        
        # Verificar el contenido del archivo Excel
        wb = load_workbook(self.excel_file)
        ws = wb.active
        
        # Verificar el encabezado
        self.assertTrue(ws.cell(row=1, column=1).value == "Asignatura: PFIS", "El encabezado de la asignatura no coincide")
        self.assertTrue(ws.cell(row=2, column=1).value == f"Fecha: {datetime.now().strftime('%d/%m/%Y')}", "La fecha no coincide")
        self.assertTrue(ws.cell(row=3, column=1).value == "Semana de Docencia: Semana 2", "La semana de docencia no coincide")
        self.assertTrue(ws.cell(row=4, column=1).value == "Tema: Tema_AlvaroMartinez", "El tema no coincide")
        
        

    def test_convertir_excel_a_pdf(self):
        """Prueba que la función convertir_excel_a_pdf crea un archivo PDF correctamente."""
        # Primero, generar el archivo Excel
        generar_hoja_asistencia(
            self.csv_file, self.excel_file, "PFIS", "Semana 2", "Tema_AlvaroMartinez"
        )
        
        # Ejecutar la función
        convertir_excel_a_pdf(self.excel_file, self.pdf_file)
        
        # Verificar que el archivo PDF se ha creado
        self.assertTrue(os.path.exists(self.pdf_file), "El archivo PDF no se ha creado")

if __name__ == "__main__":
    unittest.main()