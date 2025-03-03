import csv
from datetime import datetime
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

def generar_hoja_asistencia(csv_file, output_excel, asignatura, semana_docencia, tema):
    alumnos = []
    with open(csv_file, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        if "Nombre" not in reader.fieldnames or "Apellido(s)" not in reader.fieldnames:
            print("Formato csv incorrecto")
        for row in reader:
            alumnos.append(f"{row['Apellido(s)']}, {row['Nombre']}")
    alumnos.sort()
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"
    
    encabezado = [
        f"Asignatura: {asignatura}",
        f"Fecha: {fecha_actual}",
        f"Semana de Docencia: {semana_docencia}",
        f"Tema: {tema}"
    ]
    for i, texto in enumerate(encabezado, start=1):
        ws.cell(row=i, column=1, value=texto)
    
    fila_inicio = len(encabezado) + 2
    
    filas_por_grupo = 8

    for i, alumno in enumerate(alumnos):
        # Calcular el grupo actual
        grupo = i // (2 * filas_por_grupo)
        #AÃ±adir espacio en blanco
        if(i%16==0):
            fila_inicio+=1
        # Calcular la fila y columna dentro del grupo
        fila = fila_inicio + (i % filas_por_grupo) + (grupo * filas_por_grupo)
        columna = 1 if (i % (2 * filas_por_grupo)) < filas_por_grupo else 3
        
        # Escribir el nombre en la celda correspondiente
        ws.cell(row=fila, column=columna, value=alumno)
    
    wb.save(output_excel)

def convertir_excel_a_pdf(input_excel, output_pdf):
    wb = load_workbook(input_excel)
    ws = wb.active
    
    c = canvas.Canvas(output_pdf, pagesize=letter)
    width, height = letter
    
    y = height - 40
    for row in ws.iter_rows(values_only=True):
        text = "                         ".join(str(cell) if cell is not None else "" for cell in row)
        c.drawString(40, y, text)
        y -= 20
        if y < 40:
            c.showPage()
            y = height - 40
    
    c.save()
    print(f"PDF generado: {output_pdf}")

generar_hoja_asistencia("participantes_ficticios.csv", "hoja_asistencia.xlsx", "PFIS", "Semana 2", "Tema_AlvaroMartinez")
convertir_excel_a_pdf("hoja_asistencia.xlsx","hoja_asistencia.pdf")
